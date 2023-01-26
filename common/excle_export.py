import openpyxl


def excle_read(excle_url, sheet_name):
    # 加载工作薄
    wb = openpyxl.load_workbook(excle_url)
    # 获取sheet对象
    sheet = wb[sheet_name]
    # 获取最大行和最大列
    all_list = []
    for row in range(2, sheet.max_row+1):
        row_list=[]
        for column in range(2, sheet.max_column+1):
            row_list.append(sheet.cell(row, column).value)
        all_list.append(row_list)
    return all_list


if __name__ == '__main__':
    print(excle_read("../data/ele.xlsx", "Sheet1"))

